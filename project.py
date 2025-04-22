import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os


grade_points = {
    'O': 10,
    'E': 9,
    'A': 8,
    'B': 7,
    'C': 6,
    'D': 5,
    'F': 0
}


subject_credits = {
    'Pyhton': 4,
    'DAA': 3,
    'COA': 3,
    'DSP': 2,
    'EEC': 4,
    'EIKT': 0,
    'Python Lab': 3,
    'COA Lab': 3,
    'DAA Lab': 3,
    'Project': 3
}

def calculate_sgpa(grades):
    total_points = 0
    total_credits = 0
    for subject, grade in grades.items():
        gp = grade_points.get(grade, None)
        if gp is None:
            return None
        credit = subject_credits[subject]
        total_points += gp * credit
        total_credits += credit
    if total_credits == 0:
        return 0
    return round(total_points / total_credits, 2)

def save_to_excel(data, sgpa):
    file = "students_sgpa.xlsx"
    if os.path.exists(file):
        wb = load_workbook(file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Reg No", "Branch"] + list(subject_credits.keys()) + ["SGPA"])

    row = [data["name"], data["reg_no"], data["branch"]] + [data["grades"].get(subj, "") for subj in subject_credits] + [sgpa]
    ws.append(row)
    wb.save(file)

def submit():
    name = name_entry.get().strip()
    reg_no = reg_no_entry.get().strip()
    branch = branch_entry.get().strip()

    if not name or not reg_no or not branch:
        messagebox.showerror("Error", "Please fill in all personal details.")
        return

    grades = {}
    for subj, entry in grade_entries.items():
        grade = entry.get().strip().upper()
        if grade not in grade_points:
            messagebox.showerror("Error", f"Invalid grade '{grade}' for {subj}")
            return
        grades[subj] = grade

    sgpa = calculate_sgpa(grades)
    if sgpa is None:
        messagebox.showerror("Error", "Error calculating SGPA")
        return

    data = {"name": name, "reg_no": reg_no, "branch": branch, "grades": grades}
    save_to_excel(data, sgpa)
    messagebox.showinfo("Success", f"SGPA calculated: {sgpa}\nData saved to Excel.")

def clear_inputs():
    name_entry.delete(0, tk.END)
    reg_no_entry.delete(0, tk.END)
    branch_entry.delete(0, tk.END)
    for entry in grade_entries.values():
        entry.delete(0, tk.END)



root = tk.Tk()
root.title("SGPA Calculator")
root.geometry("600x700")

tk.Label(root, text="Student SGPA Calculator", font=('Arial', 16)).pack(pady=10)

tk.Label(root, text="Name:").pack()
name_entry = tk.Entry(root)
name_entry.pack()

tk.Label(root, text="Registration No:").pack()
reg_no_entry = tk.Entry(root)
reg_no_entry.pack()

tk.Label(root, text="Branch:").pack()
branch_entry = tk.Entry(root)
branch_entry.pack()

tk.Label(root, text="Enter Grades (O, E, A, B, C, D, F)").pack(pady=10)

grade_entries = {}
for subj in subject_credits:
    tk.Label(root, text=f"{subj}:").pack()
    entry = tk.Entry(root)
    entry.pack()
    grade_entries[subj] = entry

tk.Button(root, text="Submit", command=submit, bg="green", fg="white").pack(pady=10)
tk.Button(root, text="Clear", command=clear_inputs, bg="red", fg="white").pack(pady=5)

root.mainloop()
